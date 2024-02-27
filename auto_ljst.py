import os
import re

import openpyxl
import pyautogui as gui
from time import sleep, strftime, localtime
import datetime
import pyperclip
import requests

#mbce37650
def QuickAccess(transaction_code):
    sleep(0.45)
    gui.hotkey('ctrl', 'n')
    sleep(3)
    gui.typewrite(transaction_code)
    gui.press('enter')
    sleep(2)
    pass


def PageDown():
    gui.press('pagedown')
    gui.press('space')
    gui.hotkey('ctrl', 'c')
    cv = pyperclip.paste()
    for i in range(3):
        gui.press('pagedown')
        sleep(0.1)
    gui.press('space')
    gui.hotkey('ctrl', 'c')
    while pyperclip.paste() != cv:
        cv = pyperclip.paste()
        for i in range(20):
            gui.press('pagedown')
            sleep(0.15)
        sleep(0.45)
        gui.press('space')
        gui.hotkey('ctrl', 'c')
        pass
    pass


def send_post(touser, message):  # 企业微信发送提醒（用户， 信息）
    send_url = 'http://10.10.250.30:5001/testsendmsg/' + touser
    respone = requests.post(send_url, data={'msg': message})
    pass


def get_max_row(sheet):
    i = sheet.max_row
    real_max_row = 0
    while i > 0:
        row_dict = {i.value for i in sheet[i]}
        if row_dict == {None}:
            i = i - 1
        else:
            real_max_row = i
            break
    return real_max_row


def save_excel():
    gui.hotkey('ctrl', 'shift', 'f7')
    sleep(3)
    gui.press('enter')
    sleep(6)
    gui.hotkey('ctrl', 'c')
    while pyperclip.paste() != 'EXPORT.XLSX':
        sleep(1.5)
        gui.hotkey('ctrl', 'c')

    sleep(1)
    gui.press('enter')
    sleep(1)
    gui.press('y')
    sleep(1)
    gui.press('left')
    for j in range(4):
        gui.press('enter')
        sleep(1)
    sleep(5)
    gui.hotkey('win', 'up')
    sleep(2)
    gui.click(1424, 7)
    sleep(2)
    gui.press('enter')
    pass


'''
# 仅物料信息
def read_excel():
    wb_zmm201 = openpyxl.load_workbook('save_coois/EXPORT.XLSX')  # 建立excel文件连接
    sheet_zmm201 = wb_zmm201['Sheet1']
    row = 2
    max_row = get_max_row(sheet_zmm201)

    wb_ljst = openpyxl.load_workbook('static/模板/1102-zcst-drmb.xlsx')  # 建立excel文件连接
    sheet_ljst = wb_ljst['Sheet1']
    row_ljst = 5
    while row <= max_row:
        sheet_ljst[f'B{row_ljst}'].value = sheet_zmm201[f'C{row}'].value  # 写入中功率-导入模板
        sheet_ljst[f'D{row_ljst}'].value = sheet_zmm201[f'F{row}'].value
        sheet_ljst[f'F{row_ljst}'].value = '1102'
        sheet_ljst[f'AO{row_ljst}'].value = 'PD'
        sheet_ljst[f'AP{row_ljst}'].value = '202'
        sheet_ljst[f'AR{row_ljst}'].value = 'X'
        sheet_ljst[f'AS{row_ljst}'].value = '1102-N'
        sheet_ljst[f'AT{row_ljst}'].value = 'ND'
        sheet_ljst[f'AU{row_ljst}'].value = 'EX'
        sheet_ljst[f'AW{row_ljst}'].value = 'E'
        sheet_ljst[f'AY{row_ljst}'].value = 'CB05'
        sheet_ljst[f'BB{row_ljst}'].value = '7'
        sheet_ljst[f'BF{row_ljst}'].value = '40'
        sheet_ljst[f'BG{row_ljst}'].value = '2'
        sheet_ljst[f'BH{row_ljst}'].value = '999'
        sheet_ljst[f'BI{row_ljst}'].value = '999'
        sheet_ljst[f'BQ{row_ljst}'].value = '201'
        sheet_ljst[f'BR{row_ljst}'].value = 'Z001'
        sheet_ljst[f'BS{row_ljst}'].value = 'X'
        row_ljst += 1
        row += 1
        pass
    time_end = strftime('%Y.%m.%d', localtime())
    wb_ljst.save(f'static/EXPORT/afternoon/1102-ljst-drmb{time_end}.xlsx')
    os.remove('save_coois/EXPORT.XLSX')

    if row_ljst != 5:
        zmm004()

def zmm201(Ago, Yesterday):
    QuickAccess('zmm201')
    gui.press('down')
    gui.hotkey('ctrl', 'a')
    gui.typewrite('1102')
    sleep(0.5)
    gui.press('tab')
    gui.hotkey('ctrl', 'a')
    gui.press('backspace')
    sleep(0.5)
    gui.press('tab')
    gui.press('tab')
    gui.hotkey('ctrl', 'a')
    gui.typewrite('1201')
    gui.press('tab')
    gui.typewrite('1208')
    sleep(0.5)
    for i in range(8):
        gui.press('tab')

    gui.typewrite(Ago)
    gui.press('tab')

    gui.typewrite(Yesterday)
    gui.hotkey('ctrl', 'tab')
    for i in range(4):
        gui.hotkey('shift', 'tab')
    sleep(0.5)
    gui.press('up')
    gui.press('f8')
    sleep(2)
    pyperclip.copy('a')
    gui.hotkey('ctrl', 'c')
    while len(pyperclip.paste()) != 10:
        sleep(1)
        gui.press('space')
        gui.hotkey('ctrl', 'c')

    sleep(0.5)
    gui.hotkey('ctrl', 'f5')   # 筛选
    sleep(0.5)
    for i in range(7):
        gui.press('tab')
    sleep(0.5)
    for i in range(9):
        gui.press('down')
    gui.press('f7')
    sleep(0.5)
    gui.press('tab')
    gui.press('tab')
    for i in range(4):
        gui.press('down')
    gui.press('f7')
    sleep(2)
    for i in range(11):
        gui.press('tab')
    gui.press('enter')
    sleep(2)
    gui.typewrite('1102')
    gui.press('down')
    gui.typewrite('E')
    gui.press('enter')
    sleep(0.5)

    save_excel()

    for turns in range(3):
        gui.hotkey('shift', 'f3')
    pass
'''


# 工厂全数据
def read_excel():
    wb_zmm201 = openpyxl.load_workbook('save_coois/EXPORT.XLSX')  # 建立excel文件连接
    sheet_zmm201 = wb_zmm201['Sheet1']
    row = 2
    max_row = get_max_row(sheet_zmm201)

    wb_ljst = openpyxl.load_workbook('static/模板/1102-zcst-drmb.xlsx')  # 建立excel文件连接
    sheet_ljst = wb_ljst['Sheet1']
    row_ljst = 5
    while row <= max_row:
        if sheet_zmm201[f'AB{row}'].value:      # 存储地点已维护
            if sheet_zmm201[f'AB{row}'].value == 'CB05':
                print('1 正常:', sheet_zmm201[f'C{row}'].value)
                pass
            else:
                sheet_ljst[f'B{row_ljst}'].value = sheet_zmm201[f'C{row}'].value
                sheet_ljst[f'D{row_ljst}'].value = sheet_zmm201[f'F{row}'].value
                sheet_ljst[f'F{row_ljst}'].value = '1102'
                sheet_ljst[f'AO{row_ljst}'].value = 'PD'
                sheet_ljst[f'AP{row_ljst}'].value = '202'
                sheet_ljst[f'AR{row_ljst}'].value = 'X'
                sheet_ljst[f'AS{row_ljst}'].value = '1102-N'
                sheet_ljst[f'AT{row_ljst}'].value = 'ND'
                sheet_ljst[f'AU{row_ljst}'].value = 'EX'
                sheet_ljst[f'AW{row_ljst}'].value = 'E'
                sheet_ljst[f'AY{row_ljst}'].value = 'CB05'
                sheet_ljst[f'BB{row_ljst}'].value = '7'
                sheet_ljst[f'BF{row_ljst}'].value = '60'
                sheet_ljst[f'BG{row_ljst}'].value = '2'
                sheet_ljst[f'BH{row_ljst}'].value = '999'
                sheet_ljst[f'BI{row_ljst}'].value = '999'
                sheet_ljst[f'BQ{row_ljst}'].value = '201'
                sheet_ljst[f'BR{row_ljst}'].value = 'Z001'
                sheet_ljst[f'BS{row_ljst}'].value = 'X'
                row_ljst += 1
                print('4 存储地点错误:', sheet_zmm201[f'C{row}'].value)
                pass
            pass
        else:
            sheet_ljst[f'B{row_ljst}'].value = sheet_zmm201[f'C{row}'].value   # 写入中功率-总成视图-导入模板
            sheet_ljst[f'D{row_ljst}'].value = sheet_zmm201[f'F{row}'].value
            sheet_ljst[f'F{row_ljst}'].value = '1102'
            sheet_ljst[f'AO{row_ljst}'].value = 'PD'
            sheet_ljst[f'AP{row_ljst}'].value = '202'
            sheet_ljst[f'AR{row_ljst}'].value = 'X'
            sheet_ljst[f'AS{row_ljst}'].value = '1102-N'
            sheet_ljst[f'AT{row_ljst}'].value = 'ND'
            sheet_ljst[f'AU{row_ljst}'].value = 'EX'
            sheet_ljst[f'AW{row_ljst}'].value = 'E'
            sheet_ljst[f'AY{row_ljst}'].value = 'CB05'
            sheet_ljst[f'BB{row_ljst}'].value = '7'
            sheet_ljst[f'BF{row_ljst}'].value = '60'
            sheet_ljst[f'BG{row_ljst}'].value = '2'
            sheet_ljst[f'BH{row_ljst}'].value = '999'
            sheet_ljst[f'BI{row_ljst}'].value = '999'
            sheet_ljst[f'BQ{row_ljst}'].value = '201'
            sheet_ljst[f'BR{row_ljst}'].value = 'Z001'
            sheet_ljst[f'BS{row_ljst}'].value = 'X'
            row_ljst += 1
        row += 1
        pass
    time_end = strftime('%Y.%m.%d', localtime())
    wb_ljst.save(f'static/EXPORT/afternoon/1102-ljst-drmb{time_end}.xlsx')
    os.remove('save_coois/EXPORT.XLSX')

    if row_ljst != 5:
        zmm004()


def zmm201(Ago, Yesterday):
    QuickAccess('zmm201')
    gui.press('down')
    gui.hotkey('ctrl', 'a')
    gui.typewrite('1102')
    sleep(0.5)
    gui.press('tab')
    gui.hotkey('ctrl', 'a')
    gui.press('backspace')
    sleep(0.5)
    gui.press('tab')
    gui.press('tab')
    gui.hotkey('ctrl', 'a')
    gui.typewrite('1201')
    gui.press('tab')
    gui.typewrite('1208')
    sleep(0.5)
    for i in range(8):
        gui.press('tab')

    gui.typewrite(Ago)
    gui.press('tab')

    gui.typewrite(Yesterday)
    gui.hotkey('ctrl', 'tab')
    for i in range(4):
        gui.hotkey('shift', 'tab')
    sleep(0.5)
    for i in range(4):
        gui.press('down')
    gui.press('f8')
    sleep(2)
    pyperclip.copy('a')
    gui.hotkey('ctrl', 'c')
    while len(pyperclip.paste()) != 10:
        sleep(1)
        gui.press('space')
        gui.hotkey('ctrl', 'c')

    sleep(0.5)
    gui.hotkey('ctrl', 'f9')  # 更换布局
    sleep(0.5)
    gui.press('tab')
    gui.press('down')
    gui.press('enter')
    sleep(0.5)
    gui.hotkey('ctrl', 'f5')   # 筛选 
    sleep(0.5)
    for i in range(7):
        gui.press('tab')
    sleep(0.5)
    for i in range(69):
        gui.press('down')
    gui.press('f7')
    sleep(0.5)
    gui.press('tab')
    gui.press('tab')
    gui.press('down')
    gui.press('down')
    gui.press('f7')
    sleep(0.5)
    for i in range(3):
        gui.hotkey('shift', 'tab')
    gui.press('enter')
    sleep(0.5)
    gui.typewrite('1102')
    gui.press('down')
    gui.typewrite('E')
    gui.press('enter')
    sleep(0.5)

    save_excel()

    for turns in range(3):
        gui.hotkey('shift', 'f3')
    pass


def zmm004():
    QuickAccess('zmm004')
    gui.hotkey('ctrl', 'a')
    time_end = strftime('%Y.%m.%d', localtime())
    gui.typewrite(rf'C:\Users\User\PycharmProjects\pythonProject1\static\EXPORT\afternoon\1102-ljst-drmb{time_end}.xlsx')
    for i in range(5):
        gui.press('tab')
    for i in range(2):
        gui.press('space')  # MRP视图、工作计划视图
        gui.press('tab')
    for i in range(5):
        gui.press('tab')
        sleep(0.1)
    gui.press('space')  # 仓库管理
    sleep(0.1)
    gui.press('f8')
    sleep(3)
    gui.press('left')
    sleep(0.5)
    gui.press('enter')
    sleep(3)
    ff = gui.getActiveWindowTitle()
    ret = re.findall('激活向导', ff)
    if ret != []:
        sleep(2)
        gui.press('enter')

    n = 0
    gui.hotkey('ctrl', 'c')
    while pyperclip.paste() != '@08@' and pyperclip.paste() != '@0A@':  # 执行完毕
        sleep(2)
        gui.hotkey('ctrl', 'c')
        n += 1
        if n == 200:
            gui.hotkey('ctrl', 'n')
            break

    for i in range(3):
        sleep(0.5)
        gui.hotkey('shift', 'f3')  # 点击退出
    pass


if __name__ == '__main__':
    now = datetime.datetime.now()
    began = now - datetime.timedelta(days=1)
    time_began = began.strftime("%Y.%m.%d")
    time_end = strftime('%Y.%m.%d', localtime())

    gui.click(160, 880)
    sleep(1)
    gui.click(160, 844)
    sleep(1)
    zmm201(time_began, time_end)
    read_excel()

# sleep(2)
# print(gui.getActiveWindowTitle())